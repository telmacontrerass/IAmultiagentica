"""Tests for the benchmark Excel report exporter."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ci2lab.bench.excel import write_report


def _row(
    agent: str,
    task: str,
    solved: bool,
    *,
    model: str = "m",
    status: str = "success",
    total_tokens: int | None = 1000,
    false_positive: bool = False,
    wall: float = 1.0,
    sample: int = 0,
) -> dict[str, Any]:
    return {
        "agent": agent,
        "model": model,
        "task_id": task,
        "sample": sample,
        "solved": solved,
        "status": status,
        "total_tokens": total_tokens,
        "false_positive": false_positive,
        "wall_clock_s": wall,
        "tool_violation_count": 0,
    }


def test_write_report_creates_expected_sheets(tmp_path: Path) -> None:
    out = write_report([_row("a", "t1", True)], tmp_path / "r.xlsx")
    assert out is not None and out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["README", "Agent Comparison", "Per Task x Agent", "All Runs"]


def test_write_report_returns_none_when_no_valid_rows(tmp_path: Path) -> None:
    rows = [_row("a", "t1", False, status="error"), _row("a", "t2", False, status="timeout")]
    out = write_report(rows, tmp_path / "r.xlsx")
    assert out is None
    assert not (tmp_path / "r.xlsx").exists()


def test_error_and_timeout_rows_are_excluded(tmp_path: Path) -> None:
    rows = [
        _row("a", "t1", True, sample=0),
        _row("a", "t1", False, sample=1, status="error"),
    ]
    out = write_report(rows, tmp_path / "r.xlsx")
    assert out is not None
    ws = load_workbook(out)["All Runs"]
    data = [r for r in ws.iter_rows(min_row=4, values_only=True) if r[0]]
    assert len(data) == 1  # only the valid sample survives


def test_same_agent_different_models_stay_separate(tmp_path: Path) -> None:
    rows = [
        _row("codex", "t1", True, model="gpt-5.5"),
        _row("codex", "t1", False, model="local"),
    ]
    out = write_report(rows, tmp_path / "r.xlsx")
    assert out is not None
    ws = load_workbook(out)["Agent Comparison"]
    labels = {(r[0], r[1]) for r in ws.iter_rows(min_row=4, values_only=True) if r[0] == "codex"}
    assert ("codex", "gpt-5.5") in labels
    assert ("codex", "local") in labels

"""Export benchmark results to a formatted Excel workbook.

Turns the ``results.jsonl`` rows written by ``ci2lab bench run`` into a
multi-sheet ``.xlsx`` report (README, agent comparison, per task × agent, and
one row per sample). Only **valid runs** are included — rows whose status is an
infrastructure error or timeout are dropped, so the report reflects graded work
rather than CLI misconfiguration.

Results are grouped by ``(agent, model)`` rather than agent alone, so the same
adapter run against two different models (e.g. ``codex`` on a local model for H2
and on a frontier model for H1) stays on separate rows instead of being merged.

The workbook is regenerated in full on every call; it is safe to overwrite the
same path after each run. :func:`write_report` is a no-op that returns ``None``
when there are no valid rows.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ci2lab.bench.metrics import compute_cost_usd, load_prices, mean, median, pass_at_k

__all__ = ["write_report"]

_ERROR_STATUSES = {"error", "timeout"}
_K_REPORT = 5

# Palette (hex, no leading '#').
_NAVY = "1F3864"
_BLUE = "2E5496"
_GREEN = "C6EFCE"
_GREEN_F = "006100"
_RED = "FFC7CE"
_RED_F = "9C0006"
_AMBER = "FFEB9C"
_AMBER_F = "9C6500"
_GREY = "F2F2F2"
_WHITE = "FFFFFF"

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT = Font(bold=True, color=_WHITE, size=11)
_TITLE_FONT = Font(bold=True, size=14, color=_NAVY)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def write_report(rows: list[dict[str, Any]], out_path: Path) -> Path | None:
    """Write a formatted Excel report of the valid runs in ``rows``.

    Args:
        rows: Benchmark run records (the parsed ``results.jsonl`` rows).
        out_path: Destination ``.xlsx`` path; overwritten if it exists.

    Returns:
        ``out_path`` when a workbook was written, or ``None`` when there were no
        valid rows to report.
    """
    valid = [r for r in rows if str(r.get("status", "")) not in _ERROR_STATUSES]
    if not valid:
        return None

    _fill_costs(valid)

    workbook = Workbook()
    _build_readme(workbook, valid)
    per_group = _aggregate_by_task_agent_model(valid)
    _build_agent_comparison(workbook, per_group)
    _build_per_task(workbook, per_group)
    _build_all_runs(workbook, valid)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)
    return out_path


def _fill_costs(rows: list[dict[str, Any]]) -> None:
    """(Re)derive ``cost_usd`` for each row from the current price table.

    Runs record ``cost_usd`` at execution time, which is ``0.0`` whenever the
    model was absent from ``prices.json`` back then. The report is the right
    place to impute cost from measured tokens against the up-to-date table, so
    the workbook always reflects current pricing. A row keeps its recorded value
    only when no price applies (``compute_cost_usd`` returns ``None``).
    """
    prices = load_prices()
    for row in rows:
        cost = compute_cost_usd(
            row.get("prompt_tokens"),
            row.get("completion_tokens"),
            str(row.get("model", "")),
            prices,
        )
        if cost is not None:
            row["cost_usd"] = cost


def _aggregate_by_task_agent_model(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Aggregate rows into one summary per ``(agent, model, task)`` group."""
    groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for row in rows:
        key = (
            str(row.get("agent", "")),
            str(row.get("model", "")),
            str(row.get("task_id", "")),
        )
        groups.setdefault(key, []).append(row)

    out: list[dict[str, Any]] = []
    for (agent, model, task), group in sorted(groups.items()):
        n = len(group)
        solved = sum(1 for r in group if r.get("solved") is True)
        tokens = [float(r["total_tokens"]) for r in group if _is_num(r.get("total_tokens"))]
        latencies = [float(r["wall_clock_s"]) for r in group if _is_num(r.get("wall_clock_s"))]
        costs = [float(r["cost_usd"]) for r in group if _is_num(r.get("cost_usd"))]
        out.append(
            {
                "agent": agent,
                "model": model,
                "task_id": task,
                "n": n,
                "solved": solved,
                "pass_at_1": round(solved / n, 4) if n else 0.0,
                "pass_at_k": round(pass_at_k(n, solved, min(_K_REPORT, n)), 4),
                "mean_total_tokens": _opt_round(mean(tokens), 0),
                "median_latency_s": _opt_round(median(latencies), 1),
                "mean_cost_usd": _opt_round(mean(costs), 6),
                "total_cost_usd": _opt_round(sum(costs), 6) if costs else None,
                "false_positives": sum(1 for r in group if r.get("false_positive") is True),
                "tool_violations": sum(int(r.get("tool_violation_count", 0) or 0) for r in group),
            }
        )
    return out


def _aggregate_by_agent_model(per_group: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Roll per-task groups up into one macro-averaged row per ``(agent, model)``."""
    by_key: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in per_group:
        by_key.setdefault((str(row["agent"]), str(row["model"])), []).append(row)

    out: list[dict[str, Any]] = []
    for (agent, model), task_rows in sorted(by_key.items()):
        tokens = [
            float(r["mean_total_tokens"]) for r in task_rows if _is_num(r["mean_total_tokens"])
        ]
        latencies = [
            float(r["median_latency_s"]) for r in task_rows if _is_num(r["median_latency_s"])
        ]
        totals = [float(r["total_cost_usd"]) for r in task_rows if _is_num(r.get("total_cost_usd"))]
        runs = sum(int(r["n"]) for r in task_rows)
        total_cost = sum(totals) if totals else None
        out.append(
            {
                "agent": agent,
                "model": model,
                "tasks": len(task_rows),
                "runs": runs,
                "macro_pass_at_1": round(
                    mean([float(r["pass_at_1"]) for r in task_rows]) or 0.0, 2
                ),
                "macro_pass_at_k": round(
                    mean([float(r["pass_at_k"]) for r in task_rows]) or 0.0, 2
                ),
                "false_positives": sum(int(r["false_positives"]) for r in task_rows),
                "mean_total_tokens": _opt_round(mean(tokens), 0),
                "median_latency_s": _opt_round(mean(latencies), 1),
                "mean_cost_usd": _opt_round(total_cost / runs, 6)
                if total_cost is not None and runs
                else None,
                "total_cost_usd": _opt_round(total_cost, 4) if total_cost is not None else None,
            }
        )
    return out


def _build_readme(workbook: Workbook, valid: list[dict[str, Any]]) -> None:
    """Write the README sheet describing scope and metric definitions."""
    ws = workbook.active
    ws.title = "README"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ci2lab Benchmark Report"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = "Valid runs only — infrastructure-error / timeout runs are excluded."
    ws["A2"].font = Font(italic=True, size=10, color="595959")

    models = sorted({str(r.get("model", "")) for r in valid})
    agents = sorted({str(r.get("agent", "")) for r in valid})
    sections: list[tuple[str, list[tuple[str, str]]]] = [
        (
            "Scope",
            [
                ("Valid runs", str(len(valid))),
                ("Agents", ", ".join(agents)),
                ("Models", ", ".join(models)),
            ],
        ),
        (
            "Metric definitions",
            [
                ("Pass@1", "Fraction of individual samples that solved the task."),
                ("Pass@5", "Task solved by at least one of k samples (best-of-k)."),
                ("False Pos", "Agent reported success but the hidden oracle failed."),
                ("ToolViol", "Count of tool-policy violations during the run."),
                ("Tokens / Latency", "Mean total tokens, median wall-clock seconds per run."),
                ("USD", "Imputed cost = measured tokens x prices.json rate (not an invoice)."),
            ],
        ),
        (
            "Sheets",
            [
                ("Agent Comparison", "Macro-averaged headline metrics per agent/model."),
                ("Per Task x Agent", "Every task × agent × model cell with all metrics."),
                ("All Runs", "One row per individual sample (the raw data)."),
            ],
        ),
    ]
    row = 4
    for title, items in sections:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11, color=_BLUE)
        row += 1
        for label, value in items:
            ws.cell(row=row, column=1, value=f"  {label}").font = Font(size=10)
            ws.cell(row=row, column=2, value=value).font = Font(size=10)
            row += 1
        row += 1
    _autosize(ws, {1: 22, 2: 74})


def _build_agent_comparison(workbook: Workbook, per_group: list[dict[str, Any]]) -> None:
    """Write the macro-averaged per-(agent, model) comparison sheet."""
    ws = workbook.create_sheet("Agent Comparison")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Agent Comparison — macro-averaged over tasks"
    ws["A1"].font = _TITLE_FONT
    headers = [
        "Agent",
        "Model",
        "Tasks",
        "Runs",
        "Pass@1",
        "Pass@5",
        "False Pos",
        "Mean Tokens",
        "Median Latency (s)",
        "Mean USD/run",
        "Total USD",
    ]
    _write_header(ws, 3, headers)
    row = 4
    for item in _aggregate_by_agent_model(per_group):
        values = [
            item["agent"],
            item["model"],
            item["tasks"],
            item["runs"],
            item["macro_pass_at_1"],
            item["macro_pass_at_k"],
            item["false_positives"],
            item["mean_total_tokens"],
            item["median_latency_s"],
            item["mean_cost_usd"],
            item["total_cost_usd"],
        ]
        _write_row(ws, row, values, left_cols=2)
        _paint_pass(ws.cell(row=row, column=5), float(item["macro_pass_at_1"]))
        if int(item["false_positives"]) > 0:
            _paint(ws.cell(row=row, column=7), _RED, _RED_F)
        row += 1
    ws.freeze_panes = "A4"
    _autosize(ws, {1: 16, 2: 18, 3: 8, 4: 7, 5: 9, 6: 9, 7: 11, 8: 13, 9: 18, 10: 13, 11: 11})


def _build_per_task(workbook: Workbook, per_group: list[dict[str, Any]]) -> None:
    """Write the per task × agent × model detail sheet."""
    ws = workbook.create_sheet("Per Task x Agent")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Per Task × Agent × Model"
    ws["A1"].font = _TITLE_FONT
    headers = [
        "Task",
        "Agent",
        "Model",
        "n",
        "Pass@1",
        "Pass@5",
        "Solved",
        "False Pos",
        "ToolViol",
        "Mean Tokens",
        "Median Latency (s)",
        "Mean USD",
    ]
    _write_header(ws, 3, headers)
    row = 4
    for item in sorted(per_group, key=lambda r: (r["agent"], r["model"], r["task_id"])):
        values = [
            item["task_id"],
            item["agent"],
            item["model"],
            item["n"],
            item["pass_at_1"],
            item["pass_at_k"],
            item["solved"],
            item["false_positives"],
            item["tool_violations"],
            item["mean_total_tokens"],
            item["median_latency_s"],
            item["mean_cost_usd"],
        ]
        _write_row(ws, row, values, left_cols=3)
        _paint_pass(ws.cell(row=row, column=5), float(item["pass_at_1"]))
        if int(item["false_positives"]) > 0:
            _paint(ws.cell(row=row, column=8), _RED, _RED_F)
        row += 1
    ws.freeze_panes = "A4"
    _autosize(
        ws, {1: 10, 2: 15, 3: 17, 4: 5, 5: 9, 6: 9, 7: 8, 8: 11, 9: 10, 10: 13, 11: 18, 12: 11}
    )


def _build_all_runs(workbook: Workbook, valid: list[dict[str, Any]]) -> None:
    """Write the raw one-row-per-sample sheet with an auto-filter."""
    ws = workbook.create_sheet("All Runs")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "All Valid Runs (one row per sample)"
    ws["A1"].font = _TITLE_FONT
    headers = [
        "Task",
        "Category",
        "Agent",
        "Model",
        "Sample",
        "Solved",
        "Status",
        "False Pos",
        "ToolViol",
        "Prompt Tok",
        "Completion Tok",
        "Total Tok",
        "USD",
        "Latency (s)",
        "Rounds",
        "Tool Calls",
    ]
    _write_header(ws, 3, headers)
    row = 4
    for r in sorted(
        valid,
        key=lambda x: (
            str(x.get("agent")),
            str(x.get("model")),
            str(x.get("task_id")),
            int(x.get("sample", 0)),
        ),
    ):
        values = [
            r.get("task_id"),
            r.get("category"),
            r.get("agent"),
            r.get("model"),
            r.get("sample"),
            bool(r.get("solved")),
            r.get("status"),
            bool(r.get("false_positive")),
            int(r.get("tool_violation_count", 0) or 0),
            r.get("prompt_tokens"),
            r.get("completion_tokens"),
            r.get("total_tokens"),
            _opt_round(r.get("cost_usd"), 6),
            _opt_round(r.get("wall_clock_s"), 1),
            r.get("rounds"),
            r.get("tool_calls"),
        ]
        _write_row(ws, row, values, left_cols=4)
        solved_cell = ws.cell(row=row, column=6)
        _paint(
            solved_cell,
            _GREEN if r.get("solved") else _RED,
            _GREEN_F if r.get("solved") else _RED_F,
        )
        if r.get("false_positive") is True:
            _paint(ws.cell(row=row, column=8), _RED, _RED_F)
        row += 1
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{max(row - 1, 4)}"
    _autosize(
        ws,
        {
            1: 10,
            2: 10,
            3: 14,
            4: 17,
            5: 7,
            6: 8,
            7: 14,
            8: 10,
            9: 10,
            10: 11,
            11: 14,
            12: 11,
            13: 11,
            14: 11,
            15: 8,
            16: 10,
        },
    )


def _write_header(ws: Any, row: int, headers: list[str]) -> None:
    """Write and style a header row."""
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = _HDR_FONT
        cell.fill = PatternFill("solid", fgColor=_BLUE)
        cell.alignment = _CENTER
        cell.border = _BORDER


def _write_row(ws: Any, row: int, values: list[Any], *, left_cols: int) -> None:
    """Write one data row; the first ``left_cols`` columns are left-aligned."""
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = _BORDER
        cell.alignment = _LEFT if col <= left_cols else _CENTER
    if row % 2 == 0:
        for col in range(1, len(values) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill.patternType is None:
                cell.fill = PatternFill("solid", fgColor=_GREY)


def _paint(cell: Any, fill: str, font_color: str) -> None:
    """Apply a solid fill and font colour to a single cell."""
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color=font_color)


def _paint_pass(cell: Any, value: float) -> None:
    """Colour a Pass@1 cell green / amber / red by threshold."""
    if value >= 0.9:
        _paint(cell, _GREEN, _GREEN_F)
    elif value >= 0.6:
        _paint(cell, _AMBER, _AMBER_F)
    else:
        _paint(cell, _RED, _RED_F)


def _autosize(ws: Any, widths: dict[int, int]) -> None:
    """Set fixed column widths by column index."""
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _is_num(value: Any) -> bool:
    """Whether ``value`` is a real number (not a bool)."""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _opt_round(value: Any, digits: int) -> float | None:
    """Round a numeric value to ``digits`` places, passing non-numbers through as ``None``."""
    if not _is_num(value):
        return None
    return round(float(value), digits)
